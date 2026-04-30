from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_NAMES = (
    "01_Вводные",
    "02_Допущения",
    "03_ТЭП",
    "04_Бюджет",
    "05_СМР",
    "06_ГПР",
    "07_План_продаж",
    "08_Кредит",
    "09_ДДС",
    "10_DSCR",
    "11_Экономика",
    "12_Риски",
    "13_Сценарии",
    "14_Оптимизация",
    "15_План_улучшений",
    "16_Поставки",
    "17_Свод",
)

RU_LABELS = {
    "project_name": "Название проекта",
    "city": "Город",
    "object_type": "Тип объекта",
    "object_class": "Класс объекта",
    "land_area": "Площадь участка",
    "land_cost": "Стоимость земли",
    "total_area": "Общая площадь",
    "sellable_area": "Продаваемая площадь",
    "sellable_ratio": "Доля продаваемой площади",
    "floors": "Этажность",
    "sale_price_per_m2": "Цена продажи за м²",
    "estimated_sale_price_per_m2": "Расчётная цена продажи за м²",
    "sale_price_source": "Источник цены продажи",
    "market_price_per_m2": "Рыночный ориентир",
    "break_even_price_per_m2": "Цена безубыточности",
    "target_margin_price_per_m2": "Цена для целевой маржи",
    "recommended_price_per_m2": "Рекомендованная цена продажи",
    "preliminary_recommended_price_per_m2": "Предварительная рекомендованная цена",
    "final_recommended_price_per_m2": "Финальная рекомендованная цена",
    "final_target_margin_price_per_m2": "Цена для целевой маржи по фактическим процентам",
    "price_iteration_count": "Количество итераций расчёта цены",
    "actual_total_interest_used_for_price": "Фактические проценты, использованные для расчёта цены",
    "price_gap_to_market": "Разница к рынку",
    "base_market_price_per_m2": "Базовая рыночная цена за м²",
    "object_type_price_coefficient": "Коэффициент типа объекта",
    "floors_price_coefficient": "Коэффициент этажности для цены",
    "scale_price_coefficient": "Коэффициент масштаба для цены",
    "target_margin": "Целевая маржа",
    "construction_cost_per_m2": "Стоимость строительства за м²",
    "gp_contract_price_per_m2": "Цена генподряда за м²",
    "estimated_cmr_cost_per_m2": "Расчётная себестоимость СМР за м²",
    "cmr_cost_source": "Источник себестоимости",
    "construction_price_per_m2": "Использованная себестоимость СМР за м²",
    "base_cmr_cost_per_m2": "Базовая стоимость СМР за м²",
    "city_coefficient": "Коэффициент города",
    "floors_coefficient": "Коэффициент этажности",
    "area_coefficient": "Коэффициент масштаба",
    "engineering_coefficient": "Коэффициент инженерии",
    "construction_months": "Срок строительства, мес.",
    "sales_months": "Срок продаж, мес.",
    "credit_share": "Доля кредита",
    "credit_rate": "Ставка кредита",
    "external_networks_included": "Наружные сети включены",
    "gas_only_cooking": "Газ только пищеприготовление",
    "foundation_type": "Тип фундамента",
    "has_underground_part": "Есть подземная часть",
    "sellable_finish_level": "Отделка реализуемых помещений",
    "above_ground_structures_rate_override": "Ставка надземных несущих конструкций, ₽/м²",
    "envelope_roof_walls_rate_override": "Ставка ограждающих конструкций / стен / кровли, ₽/м²",
    "design_cost_override": "Проектирование, ₽",
    "preparation_cost_override": "Подготовительные работы, ₽",
    "earthworks_rate_override": "Земляные работы, ₽/м²",
    "sellable_finish_rate_override": "Ставка отделки реализуемых помещений, ₽/м² NSA",
    "pile_foundation_rate_override": "Ставка свайного основания, ₽/м²",
    "pile_foundation_cost_override": "Сумма свайного основания, ₽",
    "pile_count": "Количество свай",
    "average_pile_depth": "Средняя глубина сваи, м",
    "pile_unit_cost": "Стоимость одной сваи, ₽",
    "grillage_rate_override": "Ростверк / оголовки, ₽/м²",
    "foundation_optimization_mode": "Режим расчёта свайного основания",
    "plumbing_rate_override": "Сантехнические системы, ₽/м²",
    "heating_rate_override": "Отопление / ИТП, ₽/м²",
    "electrical_rate_override": "Электроснабжение, ₽/м²",
    "low_voltage_rate_override": "Слаботочные системы, ₽/м²",
    "ventilation_rate_override": "Вентиляция / дымоудаление, ₽/м²",
    "pile_foundation_rate": "Ставка свайного основания",
    "sellable_finish_rate": "Ставка отделки реализуемых помещений",
    "sellable_finish_amount": "Сумма отделки реализуемых помещений",
    "sellable_finish_calculation_logic": "Логика расчёта 2.8",
    "plumbing_rate": "Ставка 2.11 сантехнических систем",
    "heating_rate": "Ставка 2.12 отопления / ИТП",
    "electrical_rate": "Ставка 2.13 электроснабжения",
    "low_voltage_rate": "Ставка 2.14 слаботочных систем",
    "ventilation_rate": "Ставка 2.15 вентиляции / дымоудаления",
    "pit_fencing_excluded": "Ограждение котлована исключено",
    "above_ground_structures_rate": "Ставка 2.5 надземных конструкций",
    "envelope_roof_walls_rate": "Ставка 2.6 ограждающих конструкций",
    "earthworks_rate": "Ставка земляных работ",
    "manual_budget_adjustments_comment": "Комментарий о ручных корректировках",
    "design_cost_amount": "Расчётная стоимость проектирования",
    "preparation_cost_amount": "Расчётная стоимость подготовительных работ",
    "earthworks_adjustment": "Корректировка земляных работ",
    "underground_part_adjustment": "Корректировка подземной части",
    "reserve": "Резерв",
    "design": "Проектирование",
    "technical_customer": "Технический заказчик",
    "general_contractor": "Генподряд",
    "landscaping": "Благоустройство",
    "external_networks": "Наружные сети",
    "field": "Параметр",
    "value": "Значение",
    "reason": "Причина",
    "source": "Источник",
    "name": "Статья",
    "amount": "Сумма",
    "formula": "Формула",
    "share": "Доля",
    "month": "Месяц",
    "weight": "Вес",
    "construction_cost": "Затраты строительства",
    "construction_costs": "Затраты строительства",
    "land_payment": "Платеж за землю",
    "accumulated": "Накопительно",
    "sold_area": "Проданная площадь",
    "revenue": "Выручка",
    "accumulated_sold_area": "Проданная площадь накопительно",
    "accumulated_revenue": "Выручка накопительно",
    "opening_balance": "Остаток кредита на начало",
    "drawdown": "Выборка кредита",
    "interest": "Проценты",
    "repayment": "Погашение кредита",
    "closing_balance": "Остаток кредита на конец",
    "sales_receipts": "Поступления от продаж",
    "land": "Земля",
    "other_expenses": "Прочие расходы",
    "operating_cashflow_before_financing": "Операционный денежный поток до финансирования",
    "credit_drawdown": "Выборка кредита",
    "credit_repayment": "Погашение кредита",
    "equity_required": "Собственные средства",
    "cumulative_equity_required": "Собственные средства накопительно",
    "net_cashflow": "Чистый денежный поток",
    "accumulated_cashflow": "Накопленный денежный поток",
    "cash_balance_after_financing": "Остаток денежных средств после финансирования",
    "debt_service": "Обслуживание долга",
    "dscr": "DSCR",
    "total_budget": "Итоговый бюджет",
    "budget_per_total_m2": "Бюджет на 1 м² общей площади",
    "budget_per_sellable_m2": "Бюджет на 1 м² продаваемой площади",
    "revenue_per_total_m2": "Выручка на 1 м² общей площади",
    "profit_before_interest": "Прибыль до процентов",
    "profit_after_interest": "Прибыль после процентов",
    "margin_before_interest": "Маржа до процентов",
    "margin_after_interest": "Маржа после процентов",
    "total_interest": "Всего процентов",
    "max_credit_balance": "Максимальный остаток кредита",
    "total_equity_required": "Потребность в собственных средствах",
    "minimum_dscr": "Минимальный DSCR",
    "minimum_dscr_after_sales_start": "Минимальный DSCR после начала продаж",
    "average_dscr_after_sales_start": "Средний DSCR после начала продаж",
    "months_below_1_2": "Месяцев с DSCR ниже 1.2",
    "roi_on_budget": "ROI на бюджет",
    "profit": "Прибыль",
    "margin": "Маржа",
    "code": "Код риска",
    "level": "Уровень",
    "title": "Риск",
    "description": "Описание",
    "recommendation": "Рекомендация",
    "scenario": "Код сценария",
    "scenario_name": "Сценарий",
    "margin_assessment": "Оценка маржи",
    "dscr_assessment": "Оценка DSCR",
    "margin_color": "Цвет оценки маржи",
    "dscr_color": "Цвет оценки DSCR",
    "market_revenue": "Выручка по рыночной цене",
    "target_profit": "Целевая прибыль по рынку",
    "allowed_total_cost_with_interest": "Допустимые затраты с процентами",
    "current_total_cost_with_interest": "Текущий бюджет с процентами",
    "required_budget_reduction_for_market_price": "Требуемое снижение бюджета для прохождения по рынку",
    "required_budget_reduction_mln_rub": "Снижение бюджета, млн ₽",
    "required_cmr_cost_per_m2_for_market_price": "Требуемая СМР за м² для рынка",
    "required_sellable_area_for_market_price": "Требуемая продаваемая площадь",
    "required_sale_price_for_target_margin": "Цена для целевой маржи",
    "gap_to_market_price": "Разница к рынку",
    "most_realistic_option": "Наиболее реалистичный рычаг",
    "recommendations": "Рекомендации",
    "project_revenue": "Выручка проекта",
    "project_cost": "Стоимость проекта",
    "cmr_total": "СМР",
    "chapter_1_total": "Глава 1",
    "chapter_2_total": "Глава 2",
    "chapter_3_total": "Глава 3",
    "margin_rub": "Маржа, ₽",
    "margin_percent": "Маржа, %",
    "cost_per_sellable_m2": "Себестоимость на 1 м² продаваемой площади",
    "average_sale_price_per_m2": "Средняя цена продажи на 1 м²",
    "peak_debt": "Пик долга",
    "accrued_interest": "Начисленные проценты",
    "minimum_dscr_for_summary": "Минимальный DSCR",
    "equity_share": "Доля собственных средств",
    "ending_debt_balance": "Остаток долга на конец модели",
    "pile_foundation_amount": "Сумма свайного основания",
    "engineering_systems_amount": "Сумма инженерных систем",
    "engineering_systems_share_of_cmr": "Доля инженерных систем в СМР",
    "adjusted_total_budget": "Итоговый бюджет после корректировок",
}

VALUE_LABELS = {
    "source": {
        "developer_assumption": "Девелоперское допущение",
        "developer_norm": "Норматив модели",
        "cost_estimator": "Оценщик себестоимости",
        "price_estimator": "Оценщик цены продажи",
        "user_input": "Ввод пользователя",
    },
    "field": {
        "project_name": "Название проекта",
        "city": "Город",
        "object_type": "Тип объекта",
        "object_class": "Класс объекта",
        "land_area": "Площадь участка",
        "land_cost": "Стоимость земли",
        "total_area": "Общая площадь",
        "sellable_area": "Продаваемая площадь",
        "floors": "Этажность",
        "sale_price_per_m2": "Цена продажи за м²",
        "estimated_sale_price_per_m2": "Расчётная цена продажи за м²",
        "base_market_price_per_m2": "Базовая рыночная цена за м²",
        "object_type_price_coefficient": "Коэффициент типа объекта",
        "floors_price_coefficient": "Коэффициент этажности для цены",
        "scale_price_coefficient": "Коэффициент масштаба для цены",
        "market_price_per_m2": "Рыночный ориентир цены продажи",
        "target_margin": "Целевая маржа",
        "construction_months": "Срок строительства, мес.",
        "sales_months": "Срок продаж, мес.",
        "credit_share": "Доля кредита",
        "credit_rate": "Ставка кредита",
        "external_networks_included": "Наружные сети включены",
        "gas_only_cooking": "Газ только пищеприготовление",
        "foundation_type": "Тип фундамента",
        "has_underground_part": "Есть подземная часть",
        "sellable_finish_level": "Отделка реализуемых помещений",
        "above_ground_structures_rate_override": "Ставка надземных несущих конструкций, ₽/м²",
        "envelope_roof_walls_rate_override": "Ставка ограждающих конструкций / стен / кровли, ₽/м²",
        "design_cost_override": "Проектирование, ₽",
        "preparation_cost_override": "Подготовительные работы, ₽",
        "earthworks_rate_override": "Земляные работы, ₽/м²",
        "sellable_finish_rate_override": "Ставка отделки реализуемых помещений, ₽/м² NSA",
        "pile_foundation_rate_override": "Ставка свайного основания, ₽/м²",
        "pile_foundation_cost_override": "Сумма свайного основания, ₽",
        "pile_count": "Количество свай",
        "average_pile_depth": "Средняя глубина сваи, м",
        "pile_unit_cost": "Стоимость одной сваи, ₽",
        "grillage_rate_override": "Ростверк / оголовки, ₽/м²",
        "foundation_optimization_mode": "Режим расчёта свайного основания",
        "plumbing_rate_override": "Сантехнические системы, ₽/м²",
        "heating_rate_override": "Отопление / ИТП, ₽/м²",
        "electrical_rate_override": "Электроснабжение, ₽/м²",
        "low_voltage_rate_override": "Слаботочные системы, ₽/м²",
        "ventilation_rate_override": "Вентиляция / дымоудаление, ₽/м²",
        "pile_foundation_rate": "Ставка свайного основания",
        "sellable_finish_rate": "Ставка отделки реализуемых помещений",
        "sellable_finish_amount": "Сумма отделки реализуемых помещений",
        "sellable_finish_calculation_logic": "Логика расчёта 2.8",
        "plumbing_rate": "Ставка 2.11 сантехнических систем",
        "heating_rate": "Ставка 2.12 отопления / ИТП",
        "electrical_rate": "Ставка 2.13 электроснабжения",
        "low_voltage_rate": "Ставка 2.14 слаботочных систем",
        "ventilation_rate": "Ставка 2.15 вентиляции / дымоудаления",
        "pit_fencing_excluded": "Ограждение котлована исключено",
        "above_ground_structures_rate": "Ставка 2.5 надземных конструкций",
        "envelope_roof_walls_rate": "Ставка 2.6 ограждающих конструкций",
        "earthworks_rate": "Ставка земляных работ",
        "manual_budget_adjustments_comment": "Комментарий о ручных корректировках",
        "design_cost_amount": "Расчётная стоимость проектирования",
        "preparation_cost_amount": "Расчётная стоимость подготовительных работ",
        "earthworks_adjustment": "Корректировка земляных работ",
        "underground_part_adjustment": "Корректировка подземной части",
        "reserve": "Резерв",
        "design": "Проектирование",
        "technical_customer": "Технический заказчик",
        "general_contractor": "Генподряд",
        "landscaping": "Благоустройство",
        "external_networks": "Наружные сети",
        "base_cmr_cost_per_m2": "Базовая стоимость СМР за м²",
        "city_coefficient": "Коэффициент города",
        "floors_coefficient": "Коэффициент этажности",
        "area_coefficient": "Коэффициент масштаба",
        "engineering_coefficient": "Коэффициент инженерии",
        "estimated_cmr_cost_per_m2": "Итоговая расчётная себестоимость СМР за м²",
    },
    "level": {
        "high": "Высокий",
        "medium": "Средний",
        "low": "Низкий",
        "ok": "Норма",
    },
    "object_class": {
        "economy": "эконом",
        "comfort": "комфорт",
        "business": "бизнес",
        "premium": "премиум",
        "standard": "стандарт",
    },
}

SCENARIO_COLUMNS = (
    "scenario",
    "scenario_name",
    "sale_price_per_m2",
    "sales_months",
    "gp_contract_price_per_m2",
    "construction_cost_per_m2",
    "credit_rate",
    "revenue",
    "total_budget",
    "profit_before_interest",
    "profit_after_interest",
    "margin_after_interest",
    "max_credit_balance",
    "total_equity_required",
    "minimum_dscr_after_sales_start",
    "months_below_1_2",
    "margin_assessment",
    "dscr_assessment",
)

SCENARIO_LABELS = {
    "max_credit_balance": "Максимальный кредит",
    "total_equity_required": "Собственные средства",
    "minimum_dscr_after_sales_start": "Minimum DSCR",
    "months_below_1_2": "Месяцев DSCR ниже 1.2",
}


def export_model_to_excel(model: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = _safe_filename(model["input"].get("project_name") or "model")
    path = output_dir / f"{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in SHEET_NAMES}

    _write_key_values(sheets["01_Вводные"], model["input"])
    _write_table(sheets["02_Допущения"], model["assumptions"])
    _write_key_values(sheets["03_ТЭП"], model["tep"])
    _write_detailed_budget(sheets["04_Бюджет"], model["detailed_budget"])
    _write_table(sheets["05_СМР"], model["cmr"]["items"])
    _write_work_schedule(sheets["06_ГПР"], model)
    _write_table(sheets["07_План_продаж"], model["sales_plan"])
    _write_table(sheets["08_Кредит"], model["credit"]["schedule"])
    _write_table(sheets["09_ДДС"], model["cashflow"])
    _write_table(sheets["10_DSCR"], model["dscr"]["schedule"])
    _write_key_values(sheets["11_Экономика"], model["economics"])
    _write_table(sheets["12_Риски"], model["risks"])
    _write_table(sheets["13_Сценарии"], model["scenarios"], columns=SCENARIO_COLUMNS, labels=SCENARIO_LABELS)
    _write_optimization(sheets["14_Оптимизация"], model["optimization"])
    _write_improvement_plan(sheets["15_План_улучшений"], model)
    _write_supply_plan(sheets["16_Поставки"], model["supply_plan"])
    _write_summary_sheet(sheets["17_Свод"], model["summary_metrics"])

    for sheet in workbook.worksheets:
        _style_sheet(sheet)
    _style_detailed_budget(sheets["04_Бюджет"])
    _style_work_schedule(sheets["06_ГПР"], model["work_schedule"])
    _style_scenario_sheet(sheets["13_Сценарии"])
    _style_improvement_sheet(sheets["15_План_улучшений"])
    workbook.save(path)
    return path


def _write_key_values(sheet: Any, values: dict[str, Any]) -> None:
    sheet.append(["Показатель", "Значение"])
    for key, value in values.items():
        if isinstance(value, list | dict):
            continue
        sheet.append([_label(key), _display_value(key, value)])


def _write_table(
    sheet: Any,
    rows: list[dict[str, Any]],
    columns: tuple[str, ...] | None = None,
    labels: dict[str, str] | None = None,
) -> None:
    if not rows:
        sheet.append(["Нет данных"])
        return
    headers = list(columns or rows[0].keys())
    sheet.append([_label(header, labels) for header in headers])
    for row in rows:
        sheet.append([_display_table_value(row, header) for header in headers])


def _write_optimization(sheet: Any, optimization: dict[str, Any]) -> None:
    _write_key_values(sheet, optimization)
    recommendations = optimization.get("recommendations") or []
    if recommendations:
        sheet.append([])
        sheet.append(["Рекомендации"])
        for recommendation in recommendations:
            sheet.append([recommendation])


def _write_detailed_budget(sheet: Any, detailed_budget: dict[str, Any]) -> None:
    columns = (
        "Глава",
        "Код",
        "Статья",
        "База",
        "Ед.",
        "Ставка",
        "Коэфф.",
        "Сумма",
        "Материалы, %",
        "Работы, %",
        "Материалы, ₽",
        "Работы, ₽",
        "Механизмы, ₽",
        "Накладные, ₽",
        "Примечание",
    )
    sheet.append(["ДЕТАЛЬНАЯ СТРУКТУРА БЮДЖЕТА"])
    sheet.append(list(columns))
    chapter_totals = {str(row["Глава"]): row for row in detailed_budget["chapter_totals"]}
    for chapter in ("1", "2", "3"):
        for row in detailed_budget["items"]:
            if str(row["Глава"]) == chapter:
                sheet.append([row.get(column) for column in columns])
        total = chapter_totals[chapter]
        sheet.append([chapter, "", total["Статья"], "", "", "", "", total["Сумма"], "", "", "", "", "", "", ""])
    sheet.append(["", "", "ИТОГО БЮДЖЕТ", "", "", "", "", detailed_budget["total_budget"], "", "", "", "", "", "", ""])


def _write_work_schedule(sheet: Any, model: dict[str, Any]) -> None:
    work_schedule = model["work_schedule"]
    summary = work_schedule["summary"]
    months = summary["construction_months"]
    sheet.append(["ГРАФИК ПРОИЗВОДСТВА РАБОТ"])
    sheet.append(["Название проекта", model["input"].get("project_name")])
    sheet.append(["Срок строительства", summary["construction_months"], "мес.", "Общий бюджет", model["budget"]["total_budget"]])
    sheet.append(["Дата старта", summary["start_date"], "Дата окончания", summary["end_date"]])
    sheet.append([])
    headers = ["Этап", "Начало, мес.", "Длительность, мес.", "Окончание, мес.", "Стоимость"]
    headers.extend(f"Месяц {month}" for month in range(1, months + 1))
    sheet.append(headers)
    for stage in work_schedule["stages"]:
        row = [
            stage["Этап"],
            stage["Начало, мес."],
            stage["Длительность, мес."],
            stage["Окончание, мес."],
            stage["Стоимость"],
        ]
        row.extend(stage["monthly_amounts"])
        sheet.append(row)
    sheet.append(["ИТОГО CAPEX В МЕСЯЦ", "", "", "", sum(work_schedule["month_totals"]), *work_schedule["month_totals"]])
    sheet.append(["НАКОПЛЕННЫЙ CAPEX", "", "", "", work_schedule["cumulative_capex"][-1], *work_schedule["cumulative_capex"]])
    sheet.append(["% ГОТОВНОСТИ", "", "", "", 1, *work_schedule["readiness"]])


def _write_supply_plan(sheet: Any, rows: list[dict[str, Any]]) -> None:
    columns = (
        "Месяц",
        "Дата потребности",
        "Бетон, м3",
        "Арматура, т",
        "Фасадные материалы, ₽",
        "Инженерное оборудование, ₽",
        "Дата заказа бетона",
        "Дата заказа арматуры",
        "Дата заказа фасада",
        "Комментарий",
    )
    sheet.append(["ГРАФИК ПОСТАВОК МАТЕРИАЛОВ"])
    sheet.append(list(columns))
    for row in rows:
        sheet.append([row.get(column) for column in columns])


def _write_summary_sheet(sheet: Any, summary_metrics: dict[str, Any]) -> None:
    sheet.append(["СВОДНЫЕ ПОКАЗАТЕЛИ"])
    sheet.append(["Показатель", "Значение", "Ед.", "Комментарий"])
    rows = (
        ("project_revenue", "руб.", "Итого выручка проекта"),
        ("project_cost", "руб.", "Итоговый бюджет проекта"),
        ("cmr_total", "руб.", "СМР"),
        ("chapter_1_total", "руб.", "Земля и участок"),
        ("chapter_2_total", "руб.", "Строительство"),
        ("chapter_3_total", "руб.", "Коммерческие и прочие расходы"),
        ("margin_rub", "руб.", "Прибыль после процентов"),
        ("margin_percent", "%", "Маржа после процентов"),
        ("cost_per_sellable_m2", "руб./м²", "Бюджет на 1 м² продаваемой площади"),
        ("average_sale_price_per_m2", "руб./м²", "Средняя цена продажи"),
        ("peak_debt", "руб.", "Максимальный остаток кредита"),
        ("accrued_interest", "руб.", "Начисленные проценты"),
        ("minimum_dscr_for_summary", "x", "Минимальный DSCR после начала продаж"),
        ("equity_share", "%", "Потребность в собственных средствах / бюджет"),
        ("ending_debt_balance", "руб.", "Остаток долга на конец модели"),
        ("pile_foundation_amount", "руб.", "Свайное основание / ростверк"),
        ("engineering_systems_amount", "руб.", "Сумма статей 2.11–2.15"),
        ("engineering_systems_share_of_cmr", "%", "Инженерные системы / СМР"),
        ("adjusted_total_budget", "руб.", "Итоговый бюджет после детальных корректировок"),
    )
    for key, unit, comment in rows:
        sheet.append([_label(key), summary_metrics.get(key), unit, comment])


def _write_improvement_plan(sheet: Any, model: dict[str, Any]) -> None:
    improvement = model["improvement_plan"]
    optimization = model["optimization"]
    economics = model["economics"]

    sheet.append(["А. Цель оптимизации"])
    sheet.append(["Показатель", "Значение"])
    for label, value in (
        ("Требуемое снижение бюджета", improvement["target_budget_reduction"]),
        ("Целевая СМР за м²", optimization["required_cmr_cost_per_m2_for_market_price"]),
        ("Нужная продаваемая площадь", optimization["required_sellable_area_for_market_price"]),
        ("Рекомендованная цена", economics["recommended_price_per_m2"]),
        ("Рыночный ориентир", economics["market_price_per_m2"]),
    ):
        sheet.append([label, value])

    sheet.append([])
    sheet.append(["Б. Потенциал экономии по статьям"])
    _append_ru_table(
        sheet,
        improvement["improvement_items"],
        (
            "Статья",
            "Потенциал экономии, ₽",
            "Потенциал экономии, %",
            "Сложность реализации",
            "Риск влияния на качество",
            "Приоритет",
            "Комментарий",
        ),
    )

    _append_text_section(sheet, "В. Планировочные улучшения", improvement["planning_improvements"])
    _append_text_section(sheet, "Г. Коммерческие улучшения", improvement["sales_improvements"])
    _append_text_section(sheet, "Д. Финансовые улучшения", improvement["financing_improvements"])
    _append_text_section(sheet, "Е. Приоритетные действия", improvement["priority_actions"], numbered=True)
    if improvement["warnings"]:
        _append_text_section(sheet, "Предупреждения", improvement["warnings"])


def _append_ru_table(sheet: Any, rows: list[dict[str, Any]], columns: tuple[str, ...]) -> None:
    if not rows:
        sheet.append(["Нет требуемой экономии"])
        return
    sheet.append(list(columns))
    for row in rows:
        sheet.append([row.get(column) for column in columns])


def _append_text_section(sheet: Any, title: str, items: list[str], numbered: bool = False) -> None:
    sheet.append([])
    sheet.append([title])
    if not items:
        sheet.append(["Нет рекомендаций"])
        return
    for index, item in enumerate(items, start=1):
        text = f"{index}. {item}" if numbered else item
        sheet.append([text])


def _style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)
    sheet.freeze_panes = "A2"


def _style_detailed_budget(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[2]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        label = str(row[2].value or "")
        if label.startswith("ИТОГО"):
            for cell in row:
                cell.fill = total_fill
                cell.font = Font(bold=True)
    for col in range(8, 15):
        for cell in sheet.iter_cols(min_col=col, max_col=col, min_row=3, max_row=sheet.max_row):
            for value_cell in cell:
                value_cell.number_format = '#,##0'
    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 12
    sheet.column_dimensions["C"].width = 42
    sheet.column_dimensions["O"].width = 46
    sheet.freeze_panes = "A3"


def _style_work_schedule(sheet: Any, work_schedule: dict[str, Any]) -> None:
    header_row = 6
    monthly_start_col = 6
    stage_palette = [
        "D9EAF7",
        "E2F0D9",
        "FFF2CC",
        "EADCF8",
        "DDEBF7",
        "FCE4D6",
        "D9EAD3",
        "E4DFEC",
        "DDEBF7",
        "E2F0D9",
        "F4CCCC",
    ]
    for cell in sheet[header_row]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, stage in enumerate(work_schedule["stages"], start=header_row + 1):
        fill = PatternFill("solid", fgColor=stage_palette[(row_index - header_row - 1) % len(stage_palette)])
        for col in range(monthly_start_col, monthly_start_col + work_schedule["summary"]["construction_months"]):
            cell = sheet.cell(row=row_index, column=col)
            if float(cell.value or 0) > 0:
                cell.fill = fill
            cell.number_format = '#,##0'
        sheet.cell(row=row_index, column=1).font = Font(bold=True)
    for row in range(sheet.max_row - 2, sheet.max_row + 1):
        for cell in sheet[row]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
            if isinstance(cell.value, float | int):
                cell.number_format = "0.0%" if row == sheet.max_row else '#,##0'
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 14 if col >= monthly_start_col else 18
    sheet.column_dimensions["A"].width = 34
    sheet.freeze_panes = "F7"


def _style_scenario_sheet(sheet: Any) -> None:
    if sheet.max_row < 2:
        return
    headers = {cell.value: cell.column for cell in sheet[1]}
    color_map = {
        "red": "FFC7CE",
        "yellow": "FFEB9C",
        "green": "C6EFCE",
        "gray": "D9E1F2",
    }
    margin_column = headers.get("Оценка маржи")
    dscr_column = headers.get("Оценка DSCR")
    for row in range(2, sheet.max_row + 1):
        margin_value = str(sheet.cell(row=row, column=margin_column).value or "") if margin_column else ""
        dscr_value = str(sheet.cell(row=row, column=dscr_column).value or "") if dscr_column else ""
        margin_color = {"плохо": "red", "средне": "yellow", "хорошо": "green"}.get(margin_value)
        dscr_color = {"риск": "red", "норма": "green", "нет данных": "gray"}.get(dscr_value)
        if margin_column and margin_color:
            sheet.cell(row=row, column=margin_column).fill = PatternFill("solid", fgColor=color_map[margin_color])
        if dscr_column and dscr_color:
            sheet.cell(row=row, column=dscr_column).fill = PatternFill("solid", fgColor=color_map[dscr_color])


def _style_improvement_sheet(sheet: Any) -> None:
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in range(1, sheet.max_row + 1):
        value = str(sheet.cell(row=row, column=1).value or "")
        if value.startswith(("А.", "Б.", "В.", "Г.", "Д.", "Е.", "Предупреждения")):
            for cell in sheet[row]:
                cell.fill = section_fill
                cell.font = Font(bold=True)


def _label(key: str, labels: dict[str, str] | None = None) -> str:
    if labels and key in labels:
        return labels[key]
    return RU_LABELS.get(key, key)


def _display_value(key: str, value: Any) -> Any:
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    return VALUE_LABELS.get(key, {}).get(value, value)


def _display_table_value(row: dict[str, Any], header: str) -> Any:
    value = row.get(header)
    if header == "value":
        value_key = str(row.get("field") or "")
        return _display_value(value_key, value)
    return _display_value(header, value)


def _safe_filename(value: str) -> str:
    safe = "".join(char if char.isalnum() else "_" for char in value.strip())
    return safe[:40] or "model"
