from __future__ import annotations

import asyncio
import json
import shutil
import uuid
from contextlib import contextmanager
from pathlib import Path

import pytest
from openpyxl import load_workbook

from backend.auth import authenticate_user, hash_password
from backend.main import OUTPUT_DIR, app, build_financial_model, generate_model, health
from backend.models import ProjectInput
from backend.project_history import PROJECTS_DIR, load_project_history, metadata_path_for_project, save_project_metadata
from backend.tools.cmr_splitter import split_cmr
from backend.tools.excel_exporter import export_model_to_excel


@contextmanager
def workspace_tmp_dir():
    path = Path("test-artifacts") / uuid.uuid4().hex
    path.mkdir(parents=True, exist_ok=False)
    try:
        yield path
    finally:
        shutil.rmtree(path, ignore_errors=True)


def cleanup_project(project_id: str | None) -> None:
    if project_id:
        shutil.rmtree(PROJECTS_DIR / project_id, ignore_errors=True)


class ASGIResponse:
    def __init__(self, status_code: int, headers: dict[str, str], content: bytes) -> None:
        self.status_code = status_code
        self.headers = headers
        self.content = content

    def json(self) -> dict | list:
        return json.loads(self.content.decode("utf-8"))


def api_request(method: str, path: str, json_body: dict | None = None) -> ASGIResponse:
    return asyncio.run(_api_request(method, path, json_body))


async def _api_request(method: str, path: str, json_body: dict | None = None) -> ASGIResponse:
    body = b""
    headers = [(b"host", b"testserver")]
    if json_body is not None:
        body = json.dumps(json_body).encode("utf-8")
        headers.append((b"content-type", b"application/json"))
    scope = {
        "type": "http",
        "asgi": {"version": "3.0", "spec_version": "2.3"},
        "http_version": "1.1",
        "method": method.upper(),
        "scheme": "http",
        "path": path,
        "raw_path": path.encode("utf-8"),
        "query_string": b"",
        "headers": headers,
        "client": ("testclient", 50000),
        "server": ("testserver", 80),
        "root_path": "",
    }
    received = False
    response_status = 500
    response_headers: dict[str, str] = {}
    response_body = bytearray()

    async def receive() -> dict:
        nonlocal received
        if received:
            return {"type": "http.disconnect"}
        received = True
        return {"type": "http.request", "body": body, "more_body": False}

    async def send(message: dict) -> None:
        nonlocal response_status, response_headers
        if message["type"] == "http.response.start":
            response_status = message["status"]
            response_headers = {
                key.decode("latin-1").lower(): value.decode("latin-1")
                for key, value in message.get("headers", [])
            }
        elif message["type"] == "http.response.body":
            response_body.extend(message.get("body", b""))

    await app(scope, receive, send)
    return ASGIResponse(response_status, response_headers, bytes(response_body))


def _minimal_model() -> dict:
    return build_financial_model(ProjectInput(project_name="Минимальный проект"))


def _model_with_market_gap() -> dict:
    return build_financial_model(
        ProjectInput(
            project_name="Разрыв к рынку",
            city="Якутск",
            object_type="Жилой дом",
            object_class="комфорт",
            total_area=10_000,
            sellable_area=7_800,
            floors=9,
            above_ground_structures_rate_override=80_000,
        )
    )


def _technology_model(**kwargs) -> dict:
    base = {
        "project_name": "Технология 6666",
        "total_area": 6_666,
        "sellable_area": 5_200,
        "foundation_type": "сваи",
        "has_underground_part": False,
        "sellable_finish_level": "черновая",
    }
    base.update(kwargs)
    return build_financial_model(ProjectInput(**base))


def _yakutsk_122_model(**kwargs) -> dict:
    base = {
        "project_name": "Многоквартирный жилой дом в 122 квартале г. Якутска",
        "city": "Якутск",
        "object_type": "Жилой дом",
        "object_class": "комфорт",
        "total_area": 10_795.3,
        "sellable_area": 7_800,
        "foundation_type": "сваи",
        "has_underground_part": False,
        "sellable_finish_level": "черновая",
    }
    base.update(kwargs)
    return build_financial_model(ProjectInput(**base))


def _detail_item(model: dict, name: str) -> dict:
    return next(row for row in model["detailed_budget"]["items"] if row["Статья"] == name)


def test_health_works() -> None:
    assert health() == {"status": "ok"}


def test_api_health_works() -> None:
    response = api_request("GET", "/api/health")
    assert response.status_code == 200
    assert response.json() == {
        "status": "ok",
        "app": "Construction Budget Agent",
        "version": "3",
    }


def test_user_can_login() -> None:
    with workspace_tmp_dir() as tmp_dir:
        users_file = tmp_dir / "users.json"
        users_file.write_text(
            json.dumps(
                {
                    "users": [
                        {
                            "login": "ivan",
                            "password_hash": hash_password("secure-password", salt="test-salt"),
                            "role": "user",
                        }
                    ]
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        user = authenticate_user("ivan", "secure-password", users_file=users_file)
        assert user == {"login": "ivan", "role": "user"}
        assert authenticate_user("ivan", "wrong-password", users_file=users_file) is None


def test_model_is_created_from_minimal_data() -> None:
    payload = generate_model(ProjectInput(project_name="Минимальный проект"))
    assert payload["status"] == "ok"
    assert payload["budget"]["total_budget"] > 0
    assert payload["economics"]["revenue"] > 0
    assert payload["estimated_cmr_cost_per_m2"] > 0
    assert payload["cmr_cost_source"] == "Расчётная себестоимость агента"
    assert payload["recommended_price_per_m2"] > 0
    assert payload["sale_price_source"] == "Расчётная цена продажи агента"
    assert payload["improvement_plan"]["priority_actions"]
    assert payload["assumptions"]


def test_excel_is_saved() -> None:
    before = set(OUTPUT_DIR.glob("*.xlsx")) if OUTPUT_DIR.exists() else set()
    payload = generate_model(ProjectInput(project_name="Excel test"))
    filename = payload["output_filename"]
    output_path = OUTPUT_DIR / filename
    project_id = (payload.get("project_metadata") or {}).get("project_id")
    try:
        assert output_path.exists()
        assert output_path.suffix == ".xlsx"
        assert project_id
    finally:
        for path in set(OUTPUT_DIR.glob("*.xlsx")) - before:
            path.unlink(missing_ok=True)
        if project_id:
            shutil.rmtree(PROJECTS_DIR / project_id, ignore_errors=True)


def test_api_generate_model_creates_project() -> None:
    response = api_request(
        "POST",
        "/api/generate-model",
        {
            "project_name": "API project",
            "city": "Якутск",
            "total_area": 6_666,
            "sellable_area": 5_200,
        },
    )
    assert response.status_code == 200
    payload = response.json()
    project_id = payload["project_id"]
    try:
        assert project_id
        assert payload["summary"]["total_budget"] > 0
        assert payload["excel_filename"].endswith(".xlsx")
        assert payload["download_url"] == f"/api/download/{payload['excel_filename']}"
        project_dir = PROJECTS_DIR / project_id
        assert (project_dir / "input.json").exists()
        assert (project_dir / "result.json").exists()
        assert (project_dir / "metadata.json").exists()
        assert (project_dir / payload["excel_filename"]).exists()
    finally:
        cleanup_project(project_id)


def test_api_projects_returns_history() -> None:
    generated = api_request("POST", "/api/generate-model", {"project_name": "History API project"})
    assert generated.status_code == 200
    project_id = generated.json()["project_id"]
    try:
        response = api_request("GET", "/api/projects")
        assert response.status_code == 200
        rows = response.json()
        assert any(row["project_id"] == project_id for row in rows)
    finally:
        cleanup_project(project_id)


def test_api_project_returns_full_result() -> None:
    generated = api_request("POST", "/api/generate-model", {"project_name": "Full result API project"})
    assert generated.status_code == 200
    project_id = generated.json()["project_id"]
    try:
        response = api_request("GET", f"/api/projects/{project_id}")
        assert response.status_code == 200
        payload = response.json()
        assert payload["project_id"] == project_id
        assert payload["budget"]["total_budget"] > 0
        assert payload["input"]["project_name"] == "Full result API project"
    finally:
        cleanup_project(project_id)


def test_api_download_returns_excel_file() -> None:
    generated = api_request("POST", "/api/generate-model", {"project_name": "Download API project"})
    assert generated.status_code == 200
    payload = generated.json()
    project_id = payload["project_id"]
    try:
        response = api_request("GET", payload["download_url"])
        assert response.status_code == 200
        assert response.content.startswith(b"PK")
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]
    finally:
        cleanup_project(project_id)


def test_project_calculation_is_saved_to_history() -> None:
    with workspace_tmp_dir() as tmp_dir:
        model = _minimal_model()
        excel_path = export_model_to_excel(model, tmp_dir)
        projects_dir = tmp_dir / "projects"
        metadata = save_project_metadata(model=model, excel_path=excel_path, username="ivan", projects_dir=projects_dir)
        metadata_path = metadata_path_for_project(metadata["project_id"], projects_dir)

        assert metadata_path.exists()
        saved_metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
        assert saved_metadata["project_id"] == metadata["project_id"]
        assert saved_metadata["user"] == "ivan"
        assert saved_metadata["project_name"] == model["input"]["project_name"]
        assert saved_metadata["total_budget"] == pytest.approx(model["budget"]["total_budget"])


def test_excel_from_history_is_available_for_download() -> None:
    with workspace_tmp_dir() as tmp_dir:
        model = _minimal_model()
        excel_path = export_model_to_excel(model, tmp_dir)
        metadata = save_project_metadata(model=model, excel_path=excel_path, username="ivan", projects_dir=tmp_dir / "projects")
        saved_excel_path = metadata["excel_path"]

        assert saved_excel_path
        with open(saved_excel_path, "rb") as handle:
            content = handle.read()
        assert len(content) > 0


def test_metadata_json_is_created_and_history_loads() -> None:
    with workspace_tmp_dir() as tmp_dir:
        model = _minimal_model()
        excel_path = export_model_to_excel(model, tmp_dir)
        projects_dir = tmp_dir / "projects"
        metadata = save_project_metadata(model=model, excel_path=excel_path, username="admin", projects_dir=projects_dir)
        history = load_project_history(projects_dir)

        assert (projects_dir / metadata["project_id"] / "metadata.json").exists()
        assert len(history) == 1
        assert history[0]["project_id"] == metadata["project_id"]


def test_gpr_sum_equals_budget() -> None:
    model = _minimal_model()
    gpr_sum = sum(row["amount"] for row in model["gpr"])
    assert gpr_sum == pytest.approx(model["budget"]["total_budget"], abs=0.05)


def test_sales_sum_equals_sellable_area() -> None:
    model = _minimal_model()
    sold_area = sum(row["sold_area"] for row in model["sales_plan"])
    assert sold_area == pytest.approx(model["input"]["sellable_area"], abs=0.001)


def test_cmr_split_sum_is_correct() -> None:
    cmr = split_cmr(100_000_000)
    assert sum(item["amount"] for item in cmr["items"]) == pytest.approx(100_000_000, abs=0.01)
    assert cmr["items"][0]["amount"] == pytest.approx(55_000_000)


def test_minimum_dscr_ignores_empty_first_month() -> None:
    model = _minimal_model()
    first_month = model["dscr"]["schedule"][0]
    assert first_month["sales_receipts"] == 0
    assert first_month["dscr"] is None
    assert model["dscr"]["minimum_dscr_after_sales_start"] is None or model["dscr"]["minimum_dscr_after_sales_start"] > 0


def test_sellable_ratio_is_calculated() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Sellable ratio",
            total_area=10_000,
            sellable_area=7_000,
        )
    )
    assert model["tep"]["sellable_ratio"] == pytest.approx(0.7)
    assert model["economics"]["sellable_ratio"] == pytest.approx(0.7)


def test_equity_required_when_credit_is_less_than_full_deficit() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Equity gap",
            total_area=10_000,
            sellable_area=7_800,
            credit_share=0.7,
        )
    )
    assert max(row["equity_required"] for row in model["cashflow"]) > 0
    assert model["economics"]["total_equity_required"] > 0


def test_scenarios_are_generated() -> None:
    model = _minimal_model()
    scenarios = {row["scenario"]: row for row in model["scenarios"]}
    assert set(scenarios) == {"base", "optimistic", "stress"}
    assert scenarios["optimistic"]["revenue"] > scenarios["base"]["revenue"]
    assert scenarios["stress"]["revenue"] < scenarios["base"]["revenue"]
    assert scenarios["stress"]["construction_cost_per_m2"] > scenarios["base"]["construction_cost_per_m2"]
    assert scenarios["stress"]["credit_rate"] == pytest.approx(model["input"]["credit_rate"] + 0.02)
    assert scenarios["stress"]["margin_assessment"] in {"плохо", "средне", "хорошо"}
    assert scenarios["stress"]["dscr_assessment"] in {"риск", "норма", "нет данных"}


def test_excel_contains_scenarios_sheet() -> None:
    model = _minimal_model()
    path = export_model_to_excel(model, OUTPUT_DIR)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        assert "13_Сценарии" in workbook.sheetnames
        sheet = workbook["13_Сценарии"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        assert headers == [
            "Код сценария",
            "Сценарий",
            "Цена продажи за м²",
            "Срок продаж, мес.",
            "Цена генподряда за м²",
            "Стоимость строительства за м²",
            "Ставка кредита",
            "Выручка",
            "Итоговый бюджет",
            "Прибыль до процентов",
            "Прибыль после процентов",
            "Маржа после процентов",
            "Максимальный кредит",
            "Собственные средства",
            "Minimum DSCR",
            "Месяцев DSCR ниже 1.2",
            "Оценка маржи",
            "Оценка DSCR",
        ]
        assert sheet.max_row == 4
    finally:
        if workbook is not None:
            workbook.close()
        path.unlink(missing_ok=True)


def test_excel_uses_russian_user_headers() -> None:
    model = _minimal_model()
    path = export_model_to_excel(model, OUTPUT_DIR)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        tep_labels = [row[0].value for row in workbook["03_ТЭП"].iter_rows(min_row=2, max_col=1)]
        assumptions_labels = [row[0].value for row in workbook["02_Допущения"].iter_rows(min_row=2, max_col=1)]
        economics_labels = [row[0].value for row in workbook["11_Экономика"].iter_rows(min_row=2, max_col=1)]
        cashflow_headers = [cell.value for cell in next(workbook["09_ДДС"].iter_rows(min_row=1, max_row=1))]
        risks_headers = [cell.value for cell in next(workbook["12_Риски"].iter_rows(min_row=1, max_row=1))]
        assert "Название проекта" in tep_labels
        assert "Доля продаваемой площади" in tep_labels
        assert "Расчётная себестоимость СМР за м²" in tep_labels
        assert "Расчётная цена продажи за м²" in tep_labels
        assert "Рыночный ориентир" in tep_labels
        assert "Бюджет на 1 м² общей площади" in tep_labels
        assert "Базовая стоимость СМР за м²" in assumptions_labels
        assert "Коэффициент города" in assumptions_labels
        assert "Базовая рыночная цена за м²" in assumptions_labels
        assert "Коэффициент типа объекта" in assumptions_labels
        assert "Прибыль после процентов" in economics_labels
        assert "Источник цены продажи" in economics_labels
        assert "Цена безубыточности" in economics_labels
        assert "Собственные средства" in cashflow_headers
        assert "Риск" in risks_headers
    finally:
        if workbook is not None:
            workbook.close()
        path.unlink(missing_ok=True)


def test_model_estimates_cost_without_manual_cmr_price() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Автооценка",
            city="Москва",
            object_type="Жилой дом",
            object_class="комфорт",
            total_area=10_000,
            sellable_area=7_800,
            floors=10,
        )
    )
    assert model["input"]["construction_cost_per_m2"] is None
    assert model["input"]["gp_contract_price_per_m2"] is None
    assert model["estimated_cmr_cost_per_m2"] > 0
    assert model["cmr_cost_source"] == "Расчётная себестоимость агента"
    assert model["budget"]["budget_source"] == "Детальный бюджет по статьям"
    assert model["budget"]["construction_price_per_m2"] > 0


def test_detailed_budget_is_source_even_when_gp_price_is_present() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Ручной генподряд",
            city="Москва",
            total_area=10_000,
            sellable_area=7_800,
            gp_contract_price_per_m2=123_000,
        )
    )
    assert model["cmr_cost_source"] == "Ручная цена генподряда"
    assert model["budget"]["budget_source"] == "Детальный бюджет по статьям"
    assert model["budget"]["total_budget"] == pytest.approx(model["detailed_budget"]["total_budget"], abs=0.05)


def test_yakutsk_city_coefficient_is_applied() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Якутск",
            city="Якутск",
            object_type="Жилой дом",
            object_class="комфорт",
            total_area=10_000,
            sellable_area=7_800,
            floors=9,
        )
    )
    assert model["cost_estimation_coefficients"]["city_coefficient"] == pytest.approx(1.12)


def test_model_estimates_sale_price_without_manual_price() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Автоцена",
            city="Якутск",
            object_type="Жилой дом",
            object_class="комфорт",
            total_area=10_000,
            sellable_area=7_800,
            floors=9,
        )
    )
    assert model["input"]["sale_price_per_m2"] == model["recommended_price_per_m2"]
    assert model["sale_price_source"] == "Расчётная цена продажи агента"
    assert model["recommended_price_per_m2"] > 0


def test_manual_sale_price_has_priority() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Ручная цена",
            city="Якутск",
            total_area=10_000,
            sellable_area=7_800,
            sale_price_per_m2=210_000,
        )
    )
    assert model["sale_price_source"] == "Ручная цена продажи"
    assert model["input"]["sale_price_per_m2"] == 210_000
    assert model["economics"]["revenue"] == pytest.approx(7_800 * 210_000)
    assert model["price_iteration_count"] == 0


def test_yakutsk_comfort_market_price_starts_from_180000() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Рынок Якутск",
            city="Якутск",
            object_type="Жилой дом",
            object_class="комфорт",
            total_area=10_000,
            sellable_area=7_800,
            floors=9,
        )
    )
    assert model["price_estimation_components"]["base_market_price_per_m2"] == 180_000
    assert model["market_price_per_m2"] == pytest.approx(180_000)


def test_price_warning_when_target_margin_price_is_above_market() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Дорогой бюджет",
            city="Новосибирск",
            object_type="Жилой дом",
            object_class="эконом",
            total_area=10_000,
            sellable_area=7_000,
            above_ground_structures_rate_override=100_000,
        )
    )
    assert model["target_margin_price_per_m2"] > model["market_price_per_m2"]
    assert model["price_estimation_warnings"]


def test_optimization_advisor_creates_recommendations() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Оптимизация",
            city="Якутск",
            object_type="Жилой дом",
            object_class="комфорт",
            total_area=10_000,
            sellable_area=7_800,
            floors=9,
        )
    )
    optimization = model["optimization"]
    assert optimization["recommendations"]
    assert optimization["required_budget_reduction_for_market_price"] >= 0
    assert optimization["required_cmr_cost_per_m2_for_market_price"] >= 0


def test_optimization_gap_when_recommended_price_is_above_market() -> None:
    model = _model_with_market_gap()
    assert model["recommended_price_per_m2"] > model["market_price_per_m2"]
    assert model["optimization"]["gap_to_market_price"] == pytest.approx(
        model["recommended_price_per_m2"] - model["market_price_per_m2"]
    )


def test_excel_displays_english_object_class_in_russian() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Класс comfort",
            city="Якутск",
            object_type="Жилой дом",
            object_class="comfort",
            total_area=10_000,
            sellable_area=7_800,
            floors=9,
        )
    )
    path = export_model_to_excel(model, OUTPUT_DIR)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        assert "14_Оптимизация" in workbook.sheetnames
        rows = {row[0]: row[1] for row in workbook["03_ТЭП"].iter_rows(min_row=2, values_only=True)}
        assert rows["Класс объекта"] == "комфорт"
    finally:
        if workbook is not None:
            workbook.close()
        path.unlink(missing_ok=True)


def test_improvement_plan_is_created() -> None:
    model = _minimal_model()
    improvement_plan = model["improvement_plan"]
    assert improvement_plan["summary"]
    assert improvement_plan["priority_actions"]
    assert improvement_plan["assumptions"]


def test_improvement_items_exist_when_budget_reduction_is_required() -> None:
    model = _model_with_market_gap()
    assert model["optimization"]["required_budget_reduction_for_market_price"] > 0
    assert model["improvement_plan"]["improvement_items"]


def test_improvement_items_sum_to_required_budget_reduction() -> None:
    model = _model_with_market_gap()
    target = model["improvement_plan"]["target_budget_reduction"]
    savings_sum = sum(item["Потенциал экономии, ₽"] for item in model["improvement_plan"]["improvement_items"])
    assert savings_sum == pytest.approx(target, abs=0.05)


def test_reserve_is_not_first_improvement_source() -> None:
    model = _model_with_market_gap()
    items = model["improvement_plan"]["improvement_items"]
    assert items[0]["Статья"] != "Резерв"
    assert all(item["Статья"] != "Резерв" for item in items)


def test_excel_contains_improvement_plan_sheet() -> None:
    model = _model_with_market_gap()
    path = export_model_to_excel(model, OUTPUT_DIR)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        assert "15_План_улучшений" in workbook.sheetnames
        first_cell = workbook["15_План_улучшений"]["A1"].value
        assert first_cell == "А. Цель оптимизации"
    finally:
        if workbook is not None:
            workbook.close()
        path.unlink(missing_ok=True)


def test_auto_price_is_consistent_between_tep_and_optimization_excel() -> None:
    model = _model_with_market_gap()
    path = export_model_to_excel(model, OUTPUT_DIR)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        tep_rows = {row[0]: row[1] for row in workbook["03_ТЭП"].iter_rows(min_row=2, values_only=True)}
        optimization_rows = {
            row[0]: row[1]
            for row in workbook["14_Оптимизация"].iter_rows(min_row=2, values_only=True)
            if row[0] is not None
        }
        assert tep_rows["Финальная рекомендованная цена"] == model["final_recommended_price_per_m2"]
        assert tep_rows["Цена для целевой маржи по фактическим процентам"] == optimization_rows["Цена для целевой маржи"]
    finally:
        if workbook is not None:
            workbook.close()
        path.unlink(missing_ok=True)


def test_manual_sale_price_is_not_overwritten_by_final_recalculation() -> None:
    model = build_financial_model(
        ProjectInput(
            project_name="Ручная цена без перезаписи",
            city="Якутск",
            total_area=10_000,
            sellable_area=7_800,
            sale_price_per_m2=210_000,
        )
    )
    assert model["input"]["sale_price_per_m2"] == 210_000
    assert model["sale_price_source"] == "Ручная цена продажи"
    assert model["price_iteration_count"] == 0


def test_price_iteration_count_is_limited() -> None:
    model = _model_with_market_gap()
    assert model["price_iteration_count"] <= 3


def test_detailed_budget_generator_creates_budget_items() -> None:
    model = _minimal_model()
    detailed_budget = model["detailed_budget"]
    assert detailed_budget["items"]
    assert {"Глава", "Код", "Статья", "Сумма"}.issubset(detailed_budget["items"][0])


def test_detailed_budget_sum_equals_total_budget() -> None:
    model = _minimal_model()
    detail_sum = sum(row["Сумма"] for row in model["detailed_budget"]["items"])
    assert detail_sum == pytest.approx(model["budget"]["total_budget"], abs=0.05)


def test_detailed_budget_contains_materials_and_works() -> None:
    model = _minimal_model()
    split = model["detailed_budget"]["split_totals"]
    assert split["materials"] > 0
    assert split["works"] > 0


def test_excel_budget_sheet_contains_detailed_structure() -> None:
    model = _minimal_model()
    path = export_model_to_excel(model, OUTPUT_DIR)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        sheet = workbook["04_Бюджет"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert headers[:15] == [
            "Глава",
            "Код",
            "Статья",
            "База",
            "Ед.",
            "Ставка",
            "Коэфф.",
            "Сумма",
            "Материалы, %",
            "Работы, %",
            "Материалы, ₽",
            "Работы, ₽",
            "Механизмы, ₽",
            "Накладные, ₽",
            "Примечание",
        ]
        labels = [row[2] for row in sheet.iter_rows(min_row=3, values_only=True) if row[2]]
        assert "ИТОГО БЮДЖЕТ" in labels
    finally:
        if workbook is not None:
            workbook.close()
        path.unlink(missing_ok=True)


def test_excel_gpr_sheet_contains_monthly_work_structure() -> None:
    model = _minimal_model()
    path = export_model_to_excel(model, OUTPUT_DIR)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        sheet = workbook["06_ГПР"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=6, max_row=6))]
        assert headers[:6] == ["Этап", "Начало, мес.", "Длительность, мес.", "Окончание, мес.", "Стоимость", "Месяц 1"]
        labels = [row[0] for row in sheet.iter_rows(values_only=True) if row[0]]
        assert "ИТОГО CAPEX В МЕСЯЦ" in labels
        assert "НАКОПЛЕННЫЙ CAPEX" in labels
        assert "% ГОТОВНОСТИ" in labels
    finally:
        if workbook is not None:
            workbook.close()
        path.unlink(missing_ok=True)


def test_excel_contains_supply_plan_sheet() -> None:
    model = _minimal_model()
    path = export_model_to_excel(model, OUTPUT_DIR)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True)
        assert "16_Поставки" in workbook.sheetnames
        sheet = workbook["16_Поставки"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
        assert headers == [
            "Месяц",
            "Дата потребности",
            "Бетон, м3",
            "Арматура, т",
            "Фасадные материалы, ₽",
            "Инженерное оборудование, ₽",
            "Дата заказа бетона",
            "Дата заказа арматуры",
            "Дата заказа фасада",
            "Комментарий",
        ]
    finally:
        if workbook is not None:
            workbook.close()
        path.unlink(missing_ok=True)


def test_design_cost_for_6666_m2_is_about_10_million() -> None:
    model = _technology_model()
    design = _detail_item(model, "Проектирование")
    assert design["Сумма"] == pytest.approx(9_999_000, abs=1)
    assert model["budget"]["design"] == pytest.approx(9_999_000, abs=1)


def test_preparation_cost_for_6666_m2_is_about_5_million() -> None:
    model = _technology_model()
    preparation = _detail_item(model, "Подготовительный период")
    assert preparation["Сумма"] == pytest.approx(4_999_500, abs=1)


def test_pile_foundation_without_underground_part_zeroes_underground_part() -> None:
    model = _technology_model()
    underground = _detail_item(model, "Устройство несущих конструкций подземной части")
    assert underground["Сумма"] == 0
    assert "подземная часть отсутствует" in underground["Примечание"].lower()


def test_pile_earthworks_are_lower_than_underground_earthworks() -> None:
    pile_model = _technology_model(foundation_type="сваи", has_underground_part=False)
    underground_model = _technology_model(foundation_type="подземная часть", has_underground_part=True)
    pile_earthworks = _detail_item(pile_model, "Земляные работы")
    underground_earthworks = _detail_item(underground_model, "Земляные работы")
    assert pile_earthworks["Сумма"] < underground_earthworks["Сумма"]
    assert pile_earthworks["Ставка"] == pytest.approx(800)


def test_rough_finish_is_cheaper_than_white_box_and_finished() -> None:
    rough = _technology_model(sellable_finish_level="черновая")
    white_box = _technology_model(sellable_finish_level="white box")
    finished = _technology_model(sellable_finish_level="чистовая")
    rough_amount = _detail_item(rough, "Отделка реализуемых площадей")["Сумма"]
    white_box_amount = _detail_item(white_box, "Отделка реализуемых площадей")["Сумма"]
    finished_amount = _detail_item(finished, "Отделка реализуемых площадей")["Сумма"]
    assert rough_amount < white_box_amount < finished_amount


def test_detailed_budget_total_still_matches_budget_after_technology_adjustments() -> None:
    model = _technology_model()
    assert model["detailed_budget"]["total_budget"] == pytest.approx(model["budget"]["total_budget"], abs=0.05)


def test_above_ground_structures_for_piles_without_underground_use_19500_rate() -> None:
    model = _yakutsk_122_model()
    item = _detail_item(model, "Устройство несущих конструкций надземной части")
    assert item["Ставка"] == pytest.approx(19_500)
    assert item["Сумма"] == pytest.approx(10_795.3 * 19_500, abs=1)
    assert item["Источник значения"] == "технологическая корректировка"


def test_envelope_roof_walls_for_economy_comfort_use_8500_rate() -> None:
    model = _yakutsk_122_model()
    item = _detail_item(model, "Ограждающие конструкции / внутренние стены / кровля")
    assert item["Ставка"] == pytest.approx(8_500)
    assert item["Сумма"] == pytest.approx(10_795.3 * 8_500, abs=1)
    assert item["Источник значения"] == "технологическая корректировка"


def test_design_for_object_under_12000_m2_is_capped_at_10_million() -> None:
    model = _yakutsk_122_model()
    design = _detail_item(model, "Проектирование")
    assert design["Сумма"] == pytest.approx(10_000_000, abs=1)
    assert model["budget"]["design"] == pytest.approx(10_000_000, abs=1)


def test_above_ground_structures_manual_override_has_priority() -> None:
    model = _yakutsk_122_model(above_ground_structures_rate_override=17_000)
    item = _detail_item(model, "Устройство несущих конструкций надземной части")
    assert item["Ставка"] == pytest.approx(17_000)
    assert item["Сумма"] == pytest.approx(10_795.3 * 17_000, abs=1)
    assert item["Источник значения"] == "ручная корректировка"


def test_envelope_roof_walls_manual_override_has_priority() -> None:
    model = _yakutsk_122_model(envelope_roof_walls_rate_override=7_000)
    item = _detail_item(model, "Ограждающие конструкции / внутренние стены / кровля")
    assert item["Ставка"] == pytest.approx(7_000)
    assert item["Сумма"] == pytest.approx(10_795.3 * 7_000, abs=1)
    assert item["Источник значения"] == "ручная корректировка"


def test_design_manual_override_has_priority() -> None:
    model = _yakutsk_122_model(design_cost_override=6_000_000)
    design = _detail_item(model, "Проектирование")
    assert design["Сумма"] == pytest.approx(6_000_000, abs=1)
    assert model["budget"]["design"] == pytest.approx(6_000_000, abs=1)
    assert design["Источник значения"] == "ручная корректировка"


def test_detailed_budget_total_matches_budget_after_key_rate_adjustments() -> None:
    model = _yakutsk_122_model(
        above_ground_structures_rate_override=17_000,
        envelope_roof_walls_rate_override=7_000,
        design_cost_override=6_000_000,
    )
    detail_sum = sum(row["Сумма"] for row in model["detailed_budget"]["items"])
    assert detail_sum == pytest.approx(model["budget"]["total_budget"], abs=0.05)


def test_optimized_pile_foundation_rate_for_piles_without_underground_is_5500() -> None:
    model = _yakutsk_122_model()
    item = _detail_item(model, "Свайное основание / ростверк")
    assert item["Ставка"] == pytest.approx(5_500)
    assert item["Сумма"] == pytest.approx(10_795.3 * 5_500, abs=1)


def test_normative_pile_foundation_rate_is_6500() -> None:
    model = _yakutsk_122_model(foundation_optimization_mode="нормативный")
    item = _detail_item(model, "Свайное основание / ростверк")
    assert item["Ставка"] == pytest.approx(6_500)
    assert item["Сумма"] == pytest.approx(10_795.3 * 6_500, abs=1)


def test_pile_foundation_cost_override_has_priority() -> None:
    model = _yakutsk_122_model(pile_foundation_cost_override=42_000_000)
    item = _detail_item(model, "Свайное основание / ростверк")
    assert item["Сумма"] == pytest.approx(42_000_000, abs=1)
    assert item["Источник значения"] == "Ручная сумма свайного основания"


def test_pile_foundation_rate_override_has_priority_over_normative() -> None:
    model = _yakutsk_122_model(pile_foundation_rate_override=4_800)
    item = _detail_item(model, "Свайное основание / ростверк")
    assert item["Ставка"] == pytest.approx(4_800)
    assert item["Сумма"] == pytest.approx(10_795.3 * 4_800, abs=1)
    assert item["Источник значения"] == "Ручная ставка свайного основания"


def test_pile_count_and_unit_cost_calculation_works() -> None:
    model = _yakutsk_122_model(pile_count=300, pile_unit_cost=120_000, grillage_rate_override=1_000)
    item = _detail_item(model, "Свайное основание / ростверк")
    expected = 300 * 120_000 + 10_795.3 * 1_000
    assert item["Сумма"] == pytest.approx(expected, abs=1)
    assert item["Источник значения"] == "Расчёт по количеству свай"


def test_engineering_default_rates_are_optimized() -> None:
    model = _yakutsk_122_model()
    assert _detail_item(model, "Сантехнические системы")["Ставка"] == pytest.approx(4_200)
    assert _detail_item(model, "Отопление / ИТП / узел учета")["Ставка"] == pytest.approx(5_200)
    assert _detail_item(model, "Электроснабжение")["Ставка"] == pytest.approx(4_600)
    assert _detail_item(model, "Слаботочные системы")["Ставка"] == pytest.approx(1_500)
    assert _detail_item(model, "Вентиляция / дымоудаление")["Ставка"] == pytest.approx(2_500)


def test_engineering_manual_overrides_have_priority() -> None:
    model = _yakutsk_122_model(
        plumbing_rate_override=3_900,
        heating_rate_override=4_900,
        electrical_rate_override=4_100,
        low_voltage_rate_override=1_200,
        ventilation_rate_override=2_100,
    )
    assert _detail_item(model, "Сантехнические системы")["Ставка"] == pytest.approx(3_900)
    assert _detail_item(model, "Отопление / ИТП / узел учета")["Ставка"] == pytest.approx(4_900)
    assert _detail_item(model, "Электроснабжение")["Ставка"] == pytest.approx(4_100)
    assert _detail_item(model, "Слаботочные системы")["Ставка"] == pytest.approx(1_200)
    assert _detail_item(model, "Вентиляция / дымоудаление")["Ставка"] == pytest.approx(2_100)


def test_lower_pile_foundation_cost_lowers_total_budget_without_redistribution() -> None:
    base = _yakutsk_122_model()
    optimized_engineering = [
        _detail_item(base, name)["Сумма"]
        for name in (
            "Сантехнические системы",
            "Отопление / ИТП / узел учета",
            "Электроснабжение",
            "Слаботочные системы",
            "Вентиляция / дымоудаление",
        )
    ]
    lowered = _yakutsk_122_model(pile_foundation_cost_override=40_000_000)
    lowered_engineering = [
        _detail_item(lowered, name)["Сумма"]
        for name in (
            "Сантехнические системы",
            "Отопление / ИТП / узел учета",
            "Электроснабжение",
            "Слаботочные системы",
            "Вентиляция / дымоудаление",
        )
    ]
    assert lowered["budget"]["total_budget"] < base["budget"]["total_budget"]
    assert lowered_engineering == pytest.approx(optimized_engineering)


def test_detailed_budget_sum_is_the_model_total_budget() -> None:
    model = _yakutsk_122_model(pile_foundation_cost_override=40_000_000, plumbing_rate_override=3_900)
    detail_sum = sum(row["Сумма"] for row in model["detailed_budget"]["items"])
    assert detail_sum == pytest.approx(model["budget"]["total_budget"], abs=0.05)


def test_rough_sellable_finish_rate_is_11450_per_nsa_m2() -> None:
    model = _yakutsk_122_model(sellable_area=7_860.32, sellable_finish_level="черновая")
    finish = _detail_item(model, "Отделка реализуемых площадей")
    assert finish["База"] == pytest.approx(7_860.32)
    assert finish["Ед."] == "м² NSA"
    assert finish["Ставка"] == pytest.approx(11_450)
    assert finish["Источник значения"] == "уровень отделки"


def test_rough_sellable_finish_amount_for_7860_32_m2_is_about_90_million() -> None:
    model = _yakutsk_122_model(sellable_area=7_860.32, sellable_finish_level="черновая")
    finish = _detail_item(model, "Отделка реализуемых площадей")
    assert finish["Сумма"] == pytest.approx(90_000_664, abs=1)


def test_sellable_finish_rate_override_has_priority() -> None:
    model = _yakutsk_122_model(sellable_area=7_860.32, sellable_finish_rate_override=9_000)
    finish = _detail_item(model, "Отделка реализуемых площадей")
    assert finish["Ставка"] == pytest.approx(9_000)
    assert finish["Сумма"] == pytest.approx(7_860.32 * 9_000, abs=1)
    assert finish["Источник значения"] == "ручная корректировка"


def test_no_sellable_finish_level_zeroes_finish_amount() -> None:
    model = _yakutsk_122_model(sellable_area=7_860.32, sellable_finish_level="без отделки")
    finish = _detail_item(model, "Отделка реализуемых площадей")
    assert finish["Ставка"] == 0
    assert finish["Сумма"] == 0


def test_detailed_budget_total_matches_budget_after_sellable_finish_change() -> None:
    model = _yakutsk_122_model(sellable_area=7_860.32, sellable_finish_level="черновая")
    detail_sum = sum(row["Сумма"] for row in model["detailed_budget"]["items"])
    assert detail_sum == pytest.approx(model["budget"]["total_budget"], abs=0.05)
